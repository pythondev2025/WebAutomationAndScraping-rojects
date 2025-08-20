from selenium import webdriver
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import time
import json
import openpyxl as op


def main():
    raw_data = extract_data()
    save_data(raw_data)


def extract_data():
    browser = webdriver.Firefox()
    browser.get("https://remoteok.com/api/")

    # html_page = browser.page_source # returns whole html page as string

    time.sleep(10)

    click_element = browser.find_element(By.ID, "rawdata-tab")
    click_element.click()

    data_str = browser.find_element(By.TAG_NAME, "pre")
    data_str = data_str.text

    raw_data = json.loads(data_str)
    raw_data = raw_data[1:]

    for jobs in raw_data:
        jobs["description"] = BeautifulSoup(jobs["description"], "html.parser").get_text(separator=" ")

    # arranged_data = json.dumps(raw_data, indent=4)

    return raw_data


def save_data(data):
    wb = op.Workbook()
    sheet = wb.active
    sheet.title = "Jobs_data"
    
    headings = list(data[0].keys())
    for heading_no in range(len(headings)):
        heading = headings[heading_no]
        sheet.cell(row=1, column=heading_no + 1).value = heading

    sheet.column_dimensions["J"].width = 150

    for i in range(len(data)):
        sheet.row_dimensions[i + 2].height = 100
        for j in range(len(headings)):
            sheet.cell(row=i + 2, column=j + 1).value = str(data[i][headings[j]])

    wb.save("Jobs_Data.xlsx")


if __name__ == "__main__":
    main()
